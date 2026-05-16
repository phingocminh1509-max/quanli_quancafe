"""
views/history_window.py
══════════════════════════════════════════════════════════════════
Quản lý Lịch Sử Giao Dịch — chuẩn KiotViet/Sapo
Layout:
  ┌─────────────────────────────────────────────────┐
  │ 📊 Dashboard mini (tổng quan theo bộ lọc)       │
  ├─────────────────────────────────────────────────┤
  │ Bộ lọc ngày nhanh + tìm kiếm + filters nâng cao│
  ├─────────────────────────────────────────────────┤
  │ Bảng hóa đơn          │ Panel chi tiết bên phải │
  ├─────────────────────────────────────────────────┤
  │ [Excel] [PDF] [Làm mới] [Đóng]  Phân trang     │
  └─────────────────────────────────────────────────┘
"""
from __future__ import annotations
from datetime import date, datetime, timedelta
from typing import Optional

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit,
    QDateEdit, QComboBox, QFrame, QWidget, QScrollArea,
    QMessageBox, QSplitter, QAbstractItemView, QSizePolicy,
    QGridLayout, QSpacerItem,
)
from PySide6.QtCore import Qt, QDate, QTimer, Signal, QThread
from PySide6.QtGui import QColor, QFont

from database.db_config import get_session
from database.models import HoaDon, ChiTietHoaDon, NhanVien, KhuyenMai

PAGE_SIZE = 30   # số hóa đơn mỗi trang

# ── Bảng màu ─────────────────────────────────────────────────────
C_BG      = "#12121E"
C_PANEL   = "#1C1C2E"
C_CARD    = "#252538"
C_BORDER  = "#2E2E45"
C_ACCENT  = "#3498DB"
C_GREEN   = "#2ECC71"
C_ORANGE  = "#E67E22"
C_RED     = "#E74C3C"
C_YELLOW  = "#F1C40F"
C_TEXT    = "#E0E0EE"
C_MUTED   = "#7070A0"

STYLE = f"""
QDialog, QWidget   {{ background:{C_BG}; color:{C_TEXT}; }}
QFrame             {{ background:{C_PANEL}; border-radius:10px; }}
QTableWidget       {{ background:{C_CARD}; border:none; border-radius:8px;
                      gridline-color:{C_BORDER}; color:{C_TEXT}; font-size:13px; }}
QTableWidget::item {{ padding:7px 6px; border-bottom:1px solid {C_BORDER}; }}
QTableWidget::item:selected {{ background:{C_ACCENT}; color:white; }}
QHeaderView::section {{ background:{C_PANEL}; color:{C_MUTED}; padding:9px 6px;
    border:none; font-weight:bold; font-size:12px; border-bottom:2px solid {C_BORDER}; }}
QLineEdit, QDateEdit, QComboBox {{
    background:{C_CARD}; border:1px solid {C_BORDER}; border-radius:6px;
    padding:7px 10px; color:{C_TEXT}; font-size:13px; }}
QLineEdit:focus, QDateEdit:focus {{ border-color:{C_ACCENT}; }}
QComboBox::drop-down {{ border:none; width:20px; }}
QComboBox QAbstractItemView {{ background:{C_CARD}; color:{C_TEXT};
    selection-background-color:{C_ACCENT}; border:1px solid {C_BORDER}; }}
QScrollBar:vertical {{ background:{C_BG}; width:6px; border-radius:3px; }}
QScrollBar::handle:vertical {{ background:{C_BORDER}; border-radius:3px; }}
"""


def _lbl(text="", color=C_TEXT, size=13, bold=False, align=Qt.AlignLeft):
    l = QLabel(text)
    w = "bold" if bold else "normal"
    l.setStyleSheet(
        f"color:{color};font-size:{size}px;font-weight:{w};"
        "background:transparent;border:none;"
    )
    l.setAlignment(align)
    l.setTextFormat(Qt.RichText)
    return l


def _btn(text, color="#34495E", h=34, w=None):
    b = QPushButton(text)
    b.setMinimumHeight(h)
    if w: b.setFixedWidth(w)
    b.setStyleSheet(
        f"background:{color};color:white;font-weight:bold;"
        f"border-radius:6px;font-size:13px;padding:0 12px;border:none;"
    )
    b.setCursor(Qt.PointingHandCursor)
    return b


def _hline():
    f = QFrame(); f.setFrameShape(QFrame.HLine)
    f.setStyleSheet(f"background:{C_BORDER};border:none;max-height:1px;")
    return f


def _card(title: str, value: str, color=C_ACCENT, icon="") -> QFrame:
    """Mini stat card cho dashboard."""
    f = QFrame()
    f.setStyleSheet(
        f"QFrame{{background:{C_CARD};border-radius:10px;"
        f"border:1px solid {C_BORDER};padding:4px;}}"
    )
    v = QVBoxLayout(f)
    v.setContentsMargins(14, 10, 14, 10)
    v.setSpacing(4)
    lbl_icon = _lbl(f"{icon} {title}", C_MUTED, 11)
    lbl_val  = _lbl(value, color, 20, True)
    lbl_val.setAlignment(Qt.AlignLeft)
    v.addWidget(lbl_icon)
    v.addWidget(lbl_val)
    return f, lbl_val   # trả về frame + label để cập nhật sau


def _tt_color(tt: str) -> str:
    return {
        "Đã thanh toán": C_GREEN,
        "Đang chờ":      C_YELLOW,
        "Đã hủy":        C_RED,
    }.get(tt, C_MUTED)


def _pttt_icon(pttt: str) -> str:
    return {"Tiền mặt": "💵", "Chuyển khoản": "📱", "Thẻ": "💳"}.get(pttt, "💰")


# ═══════════════════════════════════════════════════════════════════
# DIALOG CHÍNH
# ═══════════════════════════════════════════════════════════════════
class HistoryDialog(QDialog):
    def __init__(self, parent=None, chuc_vu: str = "Admin"):
        super().__init__(parent)
        self.setWindowTitle("📋 Quản Lý Lịch Sử Giao Dịch")
        self.resize(1280, 780)
        self.setStyleSheet(STYLE)
        self._chuc_vu  = chuc_vu
        self._page     = 0
        self._total    = 0
        self._selected_hd_id: Optional[int] = None
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(350)
        self._search_timer.timeout.connect(self._reset_and_load)

        root = QVBoxLayout(self)
        root.setContentsMargins(14, 14, 14, 14)
        root.setSpacing(10)

        # ── Tiêu đề ──────────────────────────────────────────────
        title_row = QHBoxLayout()
        title_row.addWidget(_lbl("📋  QUẢN LÝ LỊCH SỬ GIAO DỊCH", C_ACCENT, 17, True))
        title_row.addStretch()
        self.lbl_result_count = _lbl("", C_MUTED, 12)
        title_row.addWidget(self.lbl_result_count)
        root.addLayout(title_row)

        # ── Dashboard mini ────────────────────────────────────────
        dash = QHBoxLayout(); dash.setSpacing(10)
        self._card_hd,  self._val_hd  = _card("Hóa đơn",       "0",      C_ACCENT,  "🧾")
        self._card_rev, self._val_rev = _card("Doanh thu",      "0 đ",    C_GREEN,   "💰")
        self._card_disc,self._val_disc= _card("Đã giảm",        "0 đ",    C_ORANGE,  "🎉")
        self._card_avg, self._val_avg = _card("TB/đơn",         "0 đ",    C_YELLOW,  "📊")
        for f, _ in [(self._card_hd,None),(self._card_rev,None),
                     (self._card_disc,None),(self._card_avg,None)]:
            dash.addWidget(f)
        root.addLayout(dash)
        root.addWidget(_hline())

        # ── Bộ lọc ngày nhanh + tìm kiếm ────────────────────────
        filter_row1 = QHBoxLayout(); filter_row1.setSpacing(8)

        # Nút nhanh
        quick_btns = [
            ("Hôm nay",  self._f_today),
            ("Hôm qua",  self._f_yesterday),
            ("7 ngày",   self._f_7d),
            ("30 ngày",  self._f_30d),
            ("Tháng này",self._f_thismonth),
        ]
        self._quick_btns = []
        for lbl_text, fn in quick_btns:
            b = _btn(lbl_text, C_CARD, 32)
            b.clicked.connect(fn)
            filter_row1.addWidget(b)
            self._quick_btns.append(b)
        self._active_quick = None

        filter_row1.addSpacing(8)

        # Khoảng ngày thủ công
        filter_row1.addWidget(_lbl("Từ:", C_MUTED, 12))
        self.de_from = QDateEdit(QDate.currentDate())
        self.de_from.setCalendarPopup(True)
        self.de_from.setDisplayFormat("dd/MM/yyyy")
        self.de_from.setFixedWidth(120)
        self.de_from.dateChanged.connect(self._reset_and_load)
        filter_row1.addWidget(self.de_from)

        filter_row1.addWidget(_lbl("→", C_MUTED, 12))
        self.de_to = QDateEdit(QDate.currentDate())
        self.de_to.setCalendarPopup(True)
        self.de_to.setDisplayFormat("dd/MM/yyyy")
        self.de_to.setFixedWidth(120)
        self.de_to.dateChanged.connect(self._reset_and_load)
        filter_row1.addWidget(self.de_to)

        filter_row1.addSpacing(8)

        # Tìm kiếm realtime
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("🔍 Tìm mã HD, tên KH, SĐT, nhân viên...")
        self.txt_search.setMinimumWidth(260)
        self.txt_search.textChanged.connect(lambda: self._search_timer.start())
        filter_row1.addWidget(self.txt_search, stretch=1)

        root.addLayout(filter_row1)

        # ── Filters nâng cao ─────────────────────────────────────
        filter_row2 = QHBoxLayout(); filter_row2.setSpacing(8)
        filter_row2.addWidget(_lbl("Trạng thái:", C_MUTED, 12))

        self.cb_tt = QComboBox(); self.cb_tt.setFixedWidth(150)
        self.cb_tt.addItems(["Tất cả", "Đã thanh toán", "Đang chờ", "Đã hủy"])
        self.cb_tt.currentIndexChanged.connect(self._reset_and_load)
        filter_row2.addWidget(self.cb_tt)

        filter_row2.addWidget(_lbl("Thu ngân:", C_MUTED, 12))
        self.cb_nv = QComboBox(); self.cb_nv.setFixedWidth(160)
        self.cb_nv.addItem("Tất cả", None)
        self._load_nv_filter()
        self.cb_nv.currentIndexChanged.connect(self._reset_and_load)
        filter_row2.addWidget(self.cb_nv)

        filter_row2.addWidget(_lbl("Thanh toán:", C_MUTED, 12))
        self.cb_pttt = QComboBox(); self.cb_pttt.setFixedWidth(150)
        self.cb_pttt.addItems(["Tất cả", "Tiền mặt", "Chuyển khoản", "Thẻ"])
        self.cb_pttt.currentIndexChanged.connect(self._reset_and_load)
        filter_row2.addWidget(self.cb_pttt)

        filter_row2.addStretch()

        btn_refresh = _btn("🔄 Làm mới", C_ACCENT, 34)
        btn_refresh.clicked.connect(self._reset_and_load)
        filter_row2.addWidget(btn_refresh)

        btn_excel = _btn("📊 Excel", "#27AE60", 34)
        btn_excel.clicked.connect(self._export_excel)
        filter_row2.addWidget(btn_excel)

        root.addLayout(filter_row2)
        root.addWidget(_hline())

        # ── Splitter: bảng trái | chi tiết phải ─────────────────
        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(3)
        splitter.setStyleSheet(f"QSplitter::handle{{background:{C_BORDER};}}")

        # Bảng hóa đơn
        tbl_wrap = QWidget(); tbl_wrap.setStyleSheet("background:transparent;")
        tbl_v = QVBoxLayout(tbl_wrap)
        tbl_v.setContentsMargins(0, 0, 0, 0); tbl_v.setSpacing(6)

        COLS = ["#", "Mã HD", "Thời gian", "Thu ngân", "KH", "PTTT",
                "Tổng món", "Giảm", "Thực thu", "Trạng thái"]
        self.table = QTableWidget(0, len(COLS))
        self.table.setHorizontalHeaderLabels(COLS)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(3, QHeaderView.Stretch)  # Thu ngân
        hh.setSectionResizeMode(4, QHeaderView.Stretch)  # KH
        for c, w in [(0,36),(1,70),(2,120),(5,110),(6,90),(7,80),(8,95),(9,110)]:
            self.table.setColumnWidth(c, w)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.itemSelectionChanged.connect(self._on_select)
        tbl_v.addWidget(self.table)

        # Phân trang
        page_row = QHBoxLayout()
        self.btn_prev = _btn("◀ Trước", C_CARD, 30, 90)
        self.btn_prev.clicked.connect(self._prev_page)
        self.lbl_page = _lbl("Trang 1", C_MUTED, 12, align=Qt.AlignCenter)
        self.btn_next = _btn("Sau ▶", C_CARD, 30, 90)
        self.btn_next.clicked.connect(self._next_page)
        page_row.addStretch()
        page_row.addWidget(self.btn_prev)
        page_row.addWidget(self.lbl_page)
        page_row.addWidget(self.btn_next)
        page_row.addStretch()
        tbl_v.addLayout(page_row)
        splitter.addWidget(tbl_wrap)

        # Panel chi tiết bên phải
        self.detail_panel = _DetailPanel()
        splitter.addWidget(self.detail_panel)
        splitter.setSizes([820, 380])
        root.addWidget(splitter, stretch=1)

        # ── Thanh dưới ───────────────────────────────────────────
        bot = QHBoxLayout()
        btn_close = _btn("✖ Đóng", "#555577", 38, 100)
        btn_close.clicked.connect(self.accept)
        bot.addStretch()
        bot.addWidget(btn_close)
        root.addLayout(bot)

        # Mặc định: hôm nay
        self._f_today()

    # ── Load nhân viên filter ─────────────────────────────────────
    def _load_nv_filter(self):
        s = get_session()
        try:
            nvs = s.query(NhanVien).order_by(NhanVien.ten_nv).all()
            for nv in nvs:
                self.cb_nv.addItem(nv.ten_nv, nv.id)
        finally:
            s.close()

    # ── Quick date filters ────────────────────────────────────────
    def _set_date_range(self, d_from: date, d_to: date, btn_idx: int = -1):
        self.de_from.blockSignals(True); self.de_to.blockSignals(True)
        self.de_from.setDate(QDate(d_from.year, d_from.month, d_from.day))
        self.de_to.setDate(QDate(d_to.year, d_to.month, d_to.day))
        self.de_from.blockSignals(False); self.de_to.blockSignals(False)
        # Highlight active button
        for i, b in enumerate(self._quick_btns):
            active = (i == btn_idx)
            b.setStyleSheet(
                f"background:{'#2980B9' if active else C_CARD};"
                "color:white;font-weight:bold;border-radius:6px;"
                "font-size:13px;padding:0 12px;border:none;"
            )
        self._reset_and_load()

    def _f_today(self):
        t = date.today(); self._set_date_range(t, t, 0)
    def _f_yesterday(self):
        y = date.today()-timedelta(1); self._set_date_range(y, y, 1)
    def _f_7d(self):
        t = date.today(); self._set_date_range(t-timedelta(6), t, 2)
    def _f_30d(self):
        t = date.today(); self._set_date_range(t-timedelta(29), t, 3)
    def _f_thismonth(self):
        t = date.today()
        self._set_date_range(t.replace(day=1), t, 4)

    # ── Query ────────────────────────────────────────────────────
    def _build_query(self, session):
        qd_from = self.de_from.date(); qd_to = self.de_to.date()
        d_from  = date(qd_from.year(), qd_from.month(), qd_from.day())
        d_to    = date(qd_to.year(), qd_to.month(), qd_to.day())
        dt_from = datetime(d_from.year, d_from.month, d_from.day, 0, 0, 0)
        dt_to   = datetime(d_to.year,   d_to.month,   d_to.day,   23, 59, 59)

        q = (session.query(HoaDon)
             .filter(HoaDon.thoi_gian >= dt_from, HoaDon.thoi_gian <= dt_to))

        tt = self.cb_tt.currentText()
        if tt != "Tất cả":
            q = q.filter(HoaDon.trang_thai == tt)

        nv_id = self.cb_nv.currentData()
        if nv_id:
            from database.models import PhienLamViec
            q = q.join(PhienLamViec, HoaDon.ma_phien == PhienLamViec.id)\
                 .filter(PhienLamViec.ma_nv == nv_id)

        pttt = self.cb_pttt.currentText()
        if pttt != "Tất cả":
            q = q.filter(HoaDon.phuong_thuc_tt == pttt)

        kw = self.txt_search.text().strip()
        if kw:
            from sqlalchemy import or_, cast, String
            from database.models import KhachHang
            q = q.outerjoin(KhachHang, HoaDon.ma_kh == KhachHang.id)
            q = q.filter(or_(
                cast(HoaDon.id, String).contains(kw),
                KhachHang.ten_kh.ilike(f"%{kw}%"),
                KhachHang.so_dien_thoai.contains(kw),
            ))

        return q.order_by(HoaDon.thoi_gian.desc())

    def _reset_and_load(self):
        self._page = 0
        self.load_data()

    def load_data(self):
        s = get_session()
        try:
            q = self._build_query(s)
            self._total = q.count()
            rows = q.offset(self._page * PAGE_SIZE).limit(PAGE_SIZE).all()

            # Tính dashboard từ toàn bộ kết quả (không phân trang)
            all_rows  = q.all()
            total_rev = sum(r.thanh_tien or 0 for r in all_rows)
            total_disc= sum(r.giam_gia   or 0 for r in all_rows)
            avg       = total_rev / len(all_rows) if all_rows else 0

            self._val_hd.setText(f"{self._total:,}")
            self._val_rev.setText(f"{int(total_rev):,} đ")
            self._val_disc.setText(f"{int(total_disc):,} đ")
            self._val_avg.setText(f"{int(avg):,} đ")

            pages = max(1, (self._total + PAGE_SIZE - 1) // PAGE_SIZE)
            self.lbl_page.setText(f"Trang {self._page+1}/{pages}")
            self.btn_prev.setEnabled(self._page > 0)
            self.btn_next.setEnabled(self._page < pages - 1)

            kw = self.txt_search.text().strip()
            self.lbl_result_count.setText(
                f"{'Tìm thấy' if kw else 'Tổng'}: {self._total} hóa đơn"
            )

            self.table.setRowCount(0)
            for i, hd in enumerate(rows):
                r = self.table.rowCount()
                self.table.insertRow(r)
                nv_ten = "—"
                if hd.phien_lam_viec and hd.phien_lam_viec.nhan_vien:
                    nv_ten = hd.phien_lam_viec.nhan_vien.ten_nv
                kh_ten = "—"
                if hd.khach_hang:
                    kh_ten = hd.khach_hang.ten_kh

                def _it(text, color=None, align=Qt.AlignLeft, bold=False):
                    it = QTableWidgetItem(str(text))
                    it.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    it.setTextAlignment(align | Qt.AlignVCenter)
                    if color: it.setForeground(QColor(color))
                    if bold:
                        f = it.font(); f.setBold(True); it.setFont(f)
                    return it

                stt_item = _it(str(self._page * PAGE_SIZE + i + 1), C_MUTED, Qt.AlignCenter)
                stt_item.setData(Qt.UserRole, hd.id)
                self.table.setItem(r, 0, stt_item)
                self.table.setItem(r, 1, _it(f"#{hd.id}", C_ACCENT, Qt.AlignCenter, True))
                self.table.setItem(r, 2, _it(
                    hd.thoi_gian.strftime("%d/%m %H:%M") if hd.thoi_gian else "—"
                ))
                self.table.setItem(r, 3, _it(nv_ten))
                self.table.setItem(r, 4, _it(kh_ten, C_MUTED))
                self.table.setItem(r, 5, _it(
                    f"{_pttt_icon(hd.phuong_thuc_tt)} {hd.phuong_thuc_tt or '—'}"
                ))
                self.table.setItem(r, 6, _it(
                    f"{int(hd.tong_tien or 0):,}", align=Qt.AlignRight
                ))
                self.table.setItem(r, 7, _it(
                    f"-{int(hd.giam_gia or 0):,}" if hd.giam_gia else "—",
                    C_ORANGE if hd.giam_gia else C_MUTED, Qt.AlignRight
                ))
                self.table.setItem(r, 8, _it(
                    f"{int(hd.thanh_tien or 0):,}", C_GREEN, Qt.AlignRight, True
                ))
                tt = hd.trang_thai or "—"
                self.table.setItem(r, 9, _it(tt, _tt_color(tt)))
        finally:
            s.close()

    # ── Phân trang ───────────────────────────────────────────────
    def _prev_page(self):
        if self._page > 0: self._page -= 1; self.load_data()
    def _next_page(self):
        pages = max(1, (self._total + PAGE_SIZE - 1) // PAGE_SIZE)
        if self._page < pages - 1: self._page += 1; self.load_data()

    # ── Click chọn row → hiện chi tiết ──────────────────────────
    def _on_select(self):
        row = self.table.currentRow()
        if row < 0: return
        hd_id = self.table.item(row, 0).data(Qt.UserRole)
        if hd_id == self._selected_hd_id: return
        self._selected_hd_id = hd_id
        self.detail_panel.load(hd_id)

    # ── Export Excel ─────────────────────────────────────────────
    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.warning(self, "Thiếu thư viện",
                "Cần cài openpyxl:\n  pip install openpyxl"); return

        from PySide6.QtWidgets import QFileDialog
        path, _ = QFileDialog.getSaveFileName(
            self, "Lưu file Excel",
            f"LichSu_{date.today().strftime('%d%m%Y')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path: return

        s = get_session()
        try:
            rows = self._build_query(s).all()
        finally:
            s.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch sử giao dịch"
        hdr_fill = PatternFill("solid", fgColor="1A3A5C")
        hdr_font = Font(bold=True, color="FFFFFF")
        headers  = ["Mã HD", "Thời gian", "Thu ngân", "Khách hàng",
                    "PTTT", "Tổng món", "Giảm giá", "Thực thu",
                    "KM", "Trạng thái"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for row, hd in enumerate(rows, 2):
            nv_ten = "—"
            if hd.phien_lam_viec and hd.phien_lam_viec.nhan_vien:
                nv_ten = hd.phien_lam_viec.nhan_vien.ten_nv
            ws.append([
                f"#{hd.id}",
                hd.thoi_gian.strftime("%d/%m/%Y %H:%M") if hd.thoi_gian else "",
                nv_ten,
                hd.khach_hang.ten_kh if hd.khach_hang else "",
                hd.phuong_thuc_tt or "",
                hd.tong_tien or 0,
                hd.giam_gia  or 0,
                hd.thanh_tien or 0,
                hd.khuyen_mai.ten_km if hd.khuyen_mai else "",
                hd.trang_thai or "",
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        wb.save(path)
        QMessageBox.information(self, "✅ Xuất thành công",
            f"Đã lưu {len(rows)} hóa đơn\n{path}")


# ═══════════════════════════════════════════════════════════════════
# PANEL CHI TIẾT BÊN PHẢI
# ═══════════════════════════════════════════════════════════════════
class _DetailPanel(QFrame):
    def __init__(self):
        super().__init__()
        self.setStyleSheet(
            f"QFrame{{background:{C_PANEL};border-radius:10px;"
            f"border:1px solid {C_BORDER};}}"
        )
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        self.setMinimumWidth(340)

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 14, 16, 14)
        root.setSpacing(10)

        self.lbl_title = _lbl("📄 Chi tiết hóa đơn", C_ACCENT, 14, True)
        root.addWidget(self.lbl_title)
        root.addWidget(_hline())

        # Thông tin đầu
        self.info_grid = QGridLayout()
        self.info_grid.setSpacing(6)
        self._info_vals = {}
        fields = [
            ("Mã HD",    "ma_hd"),
            ("Thời gian","thoi_gian"),
            ("Thu ngân", "nhan_vien"),
            ("Khách",    "khach_hang"),
            ("PTTT",     "pttt"),
            ("KM",       "km"),
            ("Thuế",     "thue"),
            ("Giảm",     "giam"),
            ("Thực thu", "thanh_tien"),
            ("Trạng thái","trang_thai"),
        ]
        for row, (label, key) in enumerate(fields):
            lbl_k = _lbl(label + ":", C_MUTED, 12)
            lbl_v = _lbl("—", C_TEXT, 12)
            self.info_grid.addWidget(lbl_k, row, 0)
            self.info_grid.addWidget(lbl_v, row, 1)
            self._info_vals[key] = lbl_v
        root.addLayout(self.info_grid)
        root.addWidget(_hline())

        # Bảng chi tiết món
        root.addWidget(_lbl("🍽 Danh sách món", C_MUTED, 12))
        self.tbl_mon = QTableWidget(0, 4)
        self.tbl_mon.setHorizontalHeaderLabels(["Món", "SL", "Đơn giá", "T.Tiền"])
        mhh = self.tbl_mon.horizontalHeader()
        mhh.setSectionResizeMode(0, QHeaderView.Stretch)
        for c, w in [(1,30),(2,70),(3,70)]:
            self.tbl_mon.setColumnWidth(c, w)
        self.tbl_mon.verticalHeader().setVisible(False)
        self.tbl_mon.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl_mon.setSelectionMode(QAbstractItemView.NoSelection)
        self.tbl_mon.setMaximumHeight(260)
        self.tbl_mon.setStyleSheet(
            f"QTableWidget{{background:{C_CARD};border-radius:6px;"
            f"font-size:12px;gridline-color:{C_BORDER};}}"
            f"QHeaderView::section{{background:{C_PANEL};color:{C_MUTED};"
            f"font-size:11px;padding:4px;border:none;}}"
        )
        root.addWidget(self.tbl_mon)

        root.addStretch()

        # Nút hủy đơn (chỉ nếu chưa thanh toán)
        self.btn_cancel_hd = _btn("🚫 Hủy hóa đơn", C_RED, 36)
        self.btn_cancel_hd.setVisible(False)
        self.btn_cancel_hd.clicked.connect(self._cancel_order)
        root.addWidget(self.btn_cancel_hd)

        self._current_hd_id = None

    def load(self, hd_id: int):
        self._current_hd_id = hd_id
        s = get_session()
        try:
            hd = s.get(HoaDon, hd_id)
            if not hd:
                return

            nv_ten = "—"
            if hd.phien_lam_viec and hd.phien_lam_viec.nhan_vien:
                nv_ten = hd.phien_lam_viec.nhan_vien.ten_nv

            kh_ten = "—"
            if hd.khach_hang:
                kh_ten = f"{hd.khach_hang.ten_kh}"
                if hd.khach_hang.so_dien_thoai:
                    kh_ten += f"\n📱 {hd.khach_hang.so_dien_thoai}"

            km_ten = "—"
            if hd.khuyen_mai:
                km_ten = hd.khuyen_mai.ten_km

            tt = hd.trang_thai or "—"
            self.lbl_title.setText(f"📄 Hóa đơn #{hd.id}")

            def _set(key, text, color=C_TEXT):
                l = self._info_vals.get(key)
                if l:
                    l.setText(str(text))
                    l.setStyleSheet(
                        f"color:{color};font-size:12px;background:transparent;border:none;"
                    )

            _set("ma_hd",     f"#{hd.id}", C_ACCENT)
            _set("thoi_gian", hd.thoi_gian.strftime("%H:%M — %d/%m/%Y")
                              if hd.thoi_gian else "—")
            _set("nhan_vien", nv_ten)
            _set("khach_hang",kh_ten)
            _set("pttt",      f"{_pttt_icon(hd.phuong_thuc_tt)} {hd.phuong_thuc_tt or '—'}")
            _set("km",        km_ten, C_ORANGE)
            _set("thue",      f"+{int(hd.thue or 0):,} đ", C_RED)
            _set("giam",      f"-{int(hd.giam_gia or 0):,} đ", C_ORANGE)
            _set("thanh_tien",f"{int(hd.thanh_tien or 0):,} đ", C_GREEN)
            _set("trang_thai",tt, _tt_color(tt))

            # Chi tiết món
            self.tbl_mon.setRowCount(0)
            for ct in hd.chi_tiet:
                r = self.tbl_mon.rowCount()
                self.tbl_mon.insertRow(r)
                sp_ten = ct.san_pham.ten_sp if ct.san_pham else f"SP#{ct.ma_sp}"
                ghi_chu = f"\n  ↳ {ct.ghi_chu}" if ct.ghi_chu else ""
                def _mi(text, color=C_TEXT, align=Qt.AlignLeft):
                    it = QTableWidgetItem(str(text))
                    it.setFlags(Qt.ItemIsEnabled)
                    it.setTextAlignment(align | Qt.AlignVCenter)
                    it.setForeground(QColor(color))
                    return it
                self.tbl_mon.setItem(r, 0, _mi(sp_ten + ghi_chu))
                self.tbl_mon.setItem(r, 1, _mi(str(ct.so_luong), align=Qt.AlignCenter))
                self.tbl_mon.setItem(r, 2, _mi(f"{int(ct.don_gia):,}", C_MUTED, Qt.AlignRight))
                self.tbl_mon.setItem(r, 3, _mi(f"{int(ct.thanh_tien):,}", C_GREEN, Qt.AlignRight))

            self.btn_cancel_hd.setVisible(tt == "Đang chờ")
        finally:
            s.close()

    def _cancel_order(self):
        if not self._current_hd_id: return
        if QMessageBox.question(
            self, "Xác nhận hủy",
            f"Hủy hóa đơn #{self._current_hd_id}?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        ) != QMessageBox.Yes: return
        s = get_session()
        try:
            hd = s.get(HoaDon, self._current_hd_id)
            if hd: hd.trang_thai = "Đã hủy"; s.commit()
        finally: s.close()
        self.load(self._current_hd_id)